import os
import argparse
import setting as st
import setting_2 as fst
from scheduler import GradualWarmupScheduler
import torch
import torch.nn as nn
import construct_model
import numpy as np
from torch.backends import cudnn
import utils as ut
# from plot import plot_raw_MMSE
# from plot import plot_age_prediction
# from plot import plot_age_prediction_others
# from plot import plot_age_prediction_cropped_input
# from plot import plot_age_prediction_others_cropped_input
# from plot import generate_heatmap

from data_load import data_load as DL
from data_load import cwk_data_load as cDL
from data_load import jsy_data_load as jDL
from data_load import jacob_data_load as jcDL
from data_load import aal_data_load as aDL

from plot import *
from test import *
from train import *

from openpyxl import Workbook

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random

random.seed(1234)
np.random.seed(1234)
torch.manual_seed(1234)
torch.cuda.manual_seed(1234)
torch.cuda.manual_seed_all(1234)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False
# torch.backends.cudnn.benchmark = True

def main(config):
    """ 1. data preparation """
    if fst.flag_orig_npy == True:
        print('preparation of the numpy')
        if os.path.exists(st.orig_npy_dir) == False :
            os.makedirs(st.orig_npy_dir)

        """ processing """
        if st.list_data_type[st.data_type_num] == 'Density':
            cDL.Prepare_data_GM_AGE_MMSE()
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
            jDL.Prepare_data_1()
        elif st.list_data_type[st.data_type_num] == 'ADNI_Jacob_256':
            jcDL.Prepare_data_GM_WM_CSF()
        elif 'ADNI_Jacob' in st.list_data_type[st.data_type_num] :
            jcDL.Prepare_data_GM()
        elif 'ADNI_AAL_256' in st.list_data_type[st.data_type_num] :
            aDL.Prepare_data_GM()

    if fst.flag_orig_npy_other_dataset == True:
        cDL.Prepare_data_GM_age_others(dataset='ABIDE')
        cDL.Prepare_data_GM_age_others(dataset='ICBM')
        cDL.Prepare_data_GM_age_others(dataset='Cam')
        cDL.Prepare_data_GM_age_others(dataset='IXI')
        cDL.Prepare_data_GM_age_others(dataset='PPMI')

    """ 2. fold index processing """
    if fst.flag_fold_index == True:
        print('preparation of the fold index')
        if os.path.exists(st.fold_index_dir) == False:
            os.makedirs(st.fold_index_dir)
        """ save the fold index """
        ut.preparation_fold_index(config)

    """ 3. data fold preparation, orig """
    if fst.flag_orig_npy_fold == True:
        print('preparation of processed data' + '_' + st.list_data_norm_type[st.data_norm_type_num])
        if os.path.exists(st.fold_npy_dir) == False:
            os.makedirs(st.fold_npy_dir)
        for fold in range(1, config.kfold + 1):
            print("fold : " + str(fold))
            if st.list_data_norm_type[st.data_norm_type_num] == 'woNorm':
                if st.list_data_type[st.data_type_num] == 'Density':
                    DL.Prepare_fold_data_with_age_MMSE(config, fold)
                elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
                    jDL.Prepare_fold_data_1(config, fold)
                elif st.list_data_type[st.data_type_num] == 'ADNI_Jacob_256':
                    jcDL.Prepare_fold_data_1(config, fold)
                elif 'ADNI_Jacob' in st.list_data_type[st.data_type_num]:
                    aDL.Prepare_fold_data_1(config, fold)
                elif 'ADNI_AAL_256' in st.list_data_type[st.data_type_num]:
                    aDL.Prepare_fold_data_1(config, fold)

            elif st.list_data_norm_type[st.data_norm_type_num] == 'minmax':
                if st.list_data_type[st.data_type_num] == 'ADNI_JSY':
                    jDL.Prepare_fold_data_2(config, fold)
                # elif 'ADNI_AAL_256' in st.list_data_type[st.data_type_num]:
                #     aDL.Prepare_fold_data_2(config, fold)

    """ fold selection """
    start_fold = st.start_fold
    end_fold = st.end_fold

    """ ready for the directory to save result """
    list_dir_result = []
    list_wb = []
    list_ws = []
    for i in range(len(st.list_standard_eval_dir)):
        list_dir_result.append(st.dir_to_save_1 + st.list_standard_eval_dir[i])
        ut.make_dir(dir=list_dir_result[i], flag_rm=False)
        out = ut.excel_setting(start_fold=start_fold, end_fold=end_fold, result_dir=list_dir_result[i], f_name='results')
        list_wb.append(out[0])
        list_ws.append(out[1])


    ############################################################################################################
    #########################################      fold      ##################################################
    ############################################################################################################

    list_eval_metric = st.list_eval_metric
    metric_avg = [[[] for j in range(len(st.list_eval_metric))] for i in range(len(st.list_standard_eval_dir))]
    """ fold """
    for fold in range(start_fold, end_fold+1):
        print("FOLD : {}".format(fold))
        list_dir_save_model = []
        list_dir_save_model_2 = []
        list_dir_confusion = []
        list_dir_age_pred = []
        list_dir_heatmap = []
        for i in range(len(st.list_standard_eval_dir)):
            ############################################################################################################
            ########################################      directory      ###############################################
            ############################################################################################################
            """ dir to save model """
            list_dir_save_model.append(st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/weights/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_save_model[i], flag_rm=False)

            list_dir_save_model_2.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/weights_2/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_save_model_2[i], flag_rm=False)

            """ dir to save confusion matrix  """
            list_dir_confusion.append(st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/confusion')
            ut.make_dir(dir=list_dir_confusion[i], flag_rm=False)

            """ dir to save age pred """
            list_dir_age_pred.append(st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/age_pred')
            ut.make_dir(dir=list_dir_age_pred[i], flag_rm=False)

            list_dir_heatmap.append(st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/heatmap')
            ut.make_dir(dir=list_dir_heatmap[i], flag_rm=False)

        """ dir to save pyplot """
        dir_pyplot = st.dir_to_save_1 + '/pyplot/fold_{}'.format(fold)
        ut.make_dir(dir=dir_pyplot, flag_rm=False)

        """ dir to save MMSE dist """
        dir_MMSE_dist = st.dir_to_save_1 + '/MMSE_dist'
        ut.make_dir(dir=dir_MMSE_dist, flag_rm=False)


        ############################################################################################################
        ##########################################      model      #################################################
        ############################################################################################################
        """ construct model """
        model = construct_model.construct_model(config, flag_model_num=0)
        model = nn.DataParallel(model)
        ## TODO : partially load the weights from the pretrained model
        if fst.flag_classification_fine_tune== True:
            dir_to_load = st.dir_preTrain_1
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
            model_dir = ut.model_dir_to_load(fold, dir_load_model)
            model.load_state_dict(torch.load(model_dir))
            model.eval()

        elif fst.flag_classification_using_pretrained == True :
            dir_to_load = st.dir_preTrain_1
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
            model_dir = ut.model_dir_to_load(fold, dir_load_model)
            pretrained_dict = torch.load(model_dir)
            model_dict = model.state_dict()
            for k, v in pretrained_dict.items():
                if k in model_dict:
                    print(k)
            pretrained_dict = {k: v for k, v in pretrained_dict.items() if k in model_dict}
            model_dict.update(pretrained_dict)
            model.load_state_dict(model_dict)

            model_1 = construct_model.construct_model(config, flag_model_num=1)
            model_1 = nn.DataParallel(model_1)
            dir_to_load = st.dir_preTrain_1
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
            model_dir = ut.model_dir_to_load(fold, dir_load_model)
            model_1.load_state_dict(torch.load(model_dir))
            model_1.eval()

        elif fst.flag_classification_using_pretrained_2 == True :
            model_1 = construct_model.construct_model(config, flag_model_num=1)
            model_1 = nn.DataParallel(model_1)
            dir_to_load = st.dir_preTrain_1
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
            model_dir = ut.model_dir_to_load(fold, dir_load_model)
            model_1.load_state_dict(torch.load(model_dir))
            model_1.eval()

            model_2 = construct_model.construct_model(config, flag_model_num=2)
            model_2 = nn.DataParallel(model_2)
            dir_to_load = st.dir_preTrain_2
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
            model_dir = ut.model_dir_to_load(fold, dir_load_model)
            model_2.load_state_dict(torch.load(model_dir))
            model_2.eval()

        """ optimizer """
        # optimizer = torch.optim.SGD(model.parameters(), lr=config.lr, momentum=0.9, weight_decay=st.weight_decay)
        optimizer = torch.optim.Adam(model.parameters(), lr=config.lr, betas=(0.9, 0.999), eps=1e-8, weight_decay=st.weight_decay)

        scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=st.step_size, gamma=st.LR_decay_rate, last_epoch=-1)

        # scheduler_expo = torch.optim.lr_scheduler.StepLR(optimizer, step_size=st.step_size, gamma=st.LR_decay_rate, last_epoch=-1)
        # scheduler = GradualWarmupScheduler(optimizer, multiplier=1, total_epoch=10, after_scheduler=scheduler_expo)

        # scheduler_cosine = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, st.epoch)
        # scheduler = GradualWarmupScheduler(optimizer, multiplier=1, total_epoch=10, after_scheduler=scheduler_cosine)

        # scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer=optimizer, mode='min', factor=0.1, patience=10)


        ############################################################################################################
        ##########################################      training      ##############################################
        ############################################################################################################
        list_test_result = []
        """ normal classification tasks """
        if fst.flag_classification == True or fst.flag_classification_fine_tune== True:
            # train.train(config, fold, model, optimizer, scheduler, list_dir_save_model, dir_pyplot, Validation=True, Test_flag=True)
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test.test(config, fold, model, list_dir_save_model[i_tmp], list_dir_confusion[i_tmp])
                list_test_result.append(dict_test_output)
                if len(st.list_selected_for_train) == 2 and fold == 1 and st.list_standard_eval_dir[i_tmp] == '/val_auc':
                    generate_heatmap.get_multi_heatmap_2class(config, fold, model, list_dir_save_model[i_tmp], list_dir_heatmap[i_tmp])

        elif fst.flag_classification_using_pretrained == True:
            """ using pretrained patch level model """
            train_using_pretrained.train(config, fold, model, model_1, optimizer, scheduler, list_dir_save_model, dir_pyplot, Validation=True, Test_flag=True)
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test_using_pretrained.test(config, fold, model, model_1, list_dir_save_model[i_tmp], list_dir_confusion[i_tmp])
                list_test_result.append(dict_test_output)

        elif fst.flag_classification_using_pretrained_2 == True:
            """ using pretrained patch level model """
            train_using_pretrained_2.train(config, fold, model, model_1, model_2,  optimizer, scheduler, list_dir_save_model, dir_pyplot, Validation=True, Test_flag=True)
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test_using_pretrained_2.test(config, fold, model, model_1, model_2, list_dir_save_model[i_tmp], list_dir_confusion[i_tmp])
                list_test_result.append(dict_test_output)


        elif fst.flag_regression == True:
            plot_raw_MMSE.test(config, fold, dir_MMSE_dist, flag_estimation=False)
            if fst.flag_estimate_age == True:
                plot_raw_MMSE.test(config, fold, dir_MMSE_dist, flag_estimation=True)
            train_age.train(config, fold, model, optimizer, scheduler, list_dir_save_model, dir_pyplot, Validation=True, Test_flag=True)
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test_age.test(config, fold, model,list_dir_save_model[i_tmp], list_dir_confusion[i_tmp])
                list_test_result.append(dict_test_output)
                plot_age_prediction.test(config, fold, model, list_dir_save_model[i_tmp], list_dir_age_pred[i_tmp])

        elif fst.flag_multi_task == True:
            train_multi_task.train(config, fold, model, optimizer, scheduler, list_dir_save_model, dir_pyplot, Validation=True, Test_flag=True)
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test_multi_task.test(config, fold, model, list_dir_save_model[i_tmp], list_dir_confusion[i_tmp])
                list_test_result.append(dict_test_output)

        """ fill out the results on the excel sheet """
        for i_standard in range(len(st.list_standard_eval_dir)):
            for i in range(len(list_eval_metric)):
                if list_eval_metric[i] in list_test_result[i_standard]:
                    list_ws[i_standard].cell(row=2 + i + st.push_start_row, column=fold + 1, value="%.4f" % (list_test_result[i_standard][list_eval_metric[i]]))
                    metric_avg[i_standard][i].append(list_test_result[i_standard][list_eval_metric[i]])

            for i in range(len(list_eval_metric)):
                if metric_avg[i_standard][i]:
                    avg = round(np.mean(metric_avg[i_standard][i]), 4)
                    std = round(np.std(metric_avg[i_standard][i]), 4)
                    tmp = "%.4f \u00B1 %.4f" % (avg, std)
                    list_ws[i_standard].cell(row=2 + st.push_start_row + i, column=end_fold + 2, value=tmp)

            list_wb[i_standard].save(list_dir_result[i_standard] +"/results.xlsx")

    for i_standard in range(len(st.list_standard_eval_dir)):
        n_row = list_ws[i_standard].max_row
        n_col = list_ws[i_standard].max_column
        for i_row in range(1, n_row+1):
            for i_col in range(1, n_col+1):
                ca1 = list_ws[i_standard].cell(row = i_row, column = i_col)
                ca1.alignment = Alignment(horizontal='center', vertical='center')
        list_wb[i_standard].save(list_dir_result[i_standard] + "/results.xlsx")
        list_wb[i_standard].close()

    print("finished")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--kfold', type=int, default=st.kfold)
    parser.add_argument('--num_classes', type=int, default=len(st.list_selected_for_train))
    parser.add_argument('--sagital', type=int, default=st.x_size)
    parser.add_argument('--coronal', type=int, default=st.y_size)
    parser.add_argument('--axial', type=int, default=st.z_size)
    parser.add_argument('--modality', type=int, default=st.num_modality)
    parser.add_argument('--lr', type=float, default=st.lr)
    parser.add_argument('--batch_size',type=int, default=st.batch_size)
    parser.add_argument('--v_batch_size', type=int, default=st.v_batch_size)
    parser.add_argument('--num_epochs', type=int, default=st.epoch)
    parser.add_argument('--selected_model', type=str, default=st.model_name)
    # parser.add_argument('--do_training', type=bool, default=False)
    # parser.add_argument('--saveImage',type=bool, default=False)
    # parser.add_argument('--normalize',type=bool,default=True)
    # parser.add_argument('--valset',type=bool, default = True)
    # parser.add_argument('--resume', type=bool, default=True)
    config = parser.parse_args()
    main(config)

